"""Excel workbook writer for financial consolidation output."""

from decimal import Decimal
from pathlib import Path

from financial_consolidator.config import Config
from financial_consolidator.models.report import PLSummary
from financial_consolidator.models.transaction import Transaction
from financial_consolidator.utils.logging_config import get_logger
from financial_consolidator.utils.sanitize import sanitize_for_csv

logger = get_logger(__name__)

# Import openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = None  # type: ignore


class ExcelWriter:
    """Writes financial data to a multi-sheet Excel workbook.

    Generates sheets:
    - P&L Summary
    - All Transactions (Master List)
    - Deposits
    - Transfers
    - Per-account transaction history
    - Category Analysis
    - Anomalies
    """

    # Sheet name constants for formula references
    SHEET_ALL_TRANSACTIONS = "All Transactions"

    @staticmethod
    def _txn_sort_key(txn: Transaction) -> tuple:
        """Standard sort key for transactions (date, account, description).

        Used across All Transactions, Review Queue, Deposits, and Transfers sheets
        to ensure consistent row ordering.
        """
        return (txn.date, txn.account_name, txn.description)

    def __init__(self, config: Config):
        """Initialize Excel writer.

        Args:
            config: Application configuration.
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl library not installed")

        self.config = config
        self.output_config = config.output

        # Style definitions
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        self.money_positive = Font(color="006600")  # Dark green
        self.money_negative = Font(color="CC0000")  # Dark red
        self.centered = Alignment(horizontal="center")
        self.right_aligned = Alignment(horizontal="right")
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def write(
        self,
        output_path: Path,
        transactions: list[Transaction],
        date_gaps: list[dict[str, object]] | None,
        pl_summary: PLSummary,
    ) -> None:
        """Write all data to an Excel workbook.

        Args:
            output_path: Path for output file.
            transactions: List of processed transactions.
            date_gaps: Optional list of date gap anomalies.
            pl_summary: Pre-computed P&L summary data.
        """
        logger.info(f"Writing Excel workbook to {output_path}")

        wb = Workbook()
        # Remove default sheet
        if wb.active:
            wb.remove(wb.active)

        # Create sheets
        self._create_category_lookup(wb)  # Must come first for VLOOKUP references
        self._create_pl_summary(wb, pl_summary)
        self._create_master_list(wb, transactions)
        self._create_review_queue(wb, transactions)
        self._create_deposits_sheet(wb, transactions)
        self._create_transfers_sheet(wb, transactions)
        self._create_account_sheets(wb, transactions)
        self._create_account_summary(wb, transactions)
        self._create_category_analysis(wb, transactions)
        self._create_anomalies_sheet(wb, transactions, date_gaps or [])

        # Save workbook
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        logger.info(f"Excel workbook saved: {output_path}")

    def _create_category_lookup(self, wb: "Workbook") -> None:
        """Create hidden Category Lookup sheet for VLOOKUP and data validation.

        Args:
            wb: Workbook to add sheet to.
        """
        ws = wb.create_sheet("Category Lookup")

        # Headers
        headers = ["Category Name", "Category Type"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill

        # Get all categories sorted by name
        categories = sorted(
            self.config.categories.values(),
            key=lambda c: c.name
        )

        # Write category data
        row = 2
        for category in categories:
            # Skip subcategories (those with parent_id)
            if category.parent_id:
                continue

            ws.cell(row=row, column=1, value=category.name)
            cat_type = category.category_type.value if category.category_type else ""
            ws.cell(row=row, column=2, value=cat_type)
            row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15

        # Hide the sheet
        # Use veryHidden to prevent users from unhiding and deleting this sheet,
        # which would break VLOOKUP formulas in Category Type column
        ws.sheet_state = "veryHidden"

        logger.debug(f"Created Category Lookup sheet with {row - 2} categories")

    def _create_pl_summary(
        self, wb: "Workbook", pl_summary: PLSummary
    ) -> None:
        """Create P&L Summary sheet with year-by-year breakdown using SUMIFS formulas.

        Uses formulas that reference 'All Transactions' sheet so that changes
        to category assignments automatically update P&L totals.

        Args:
            wb: Workbook to add sheet to.
            pl_summary: Pre-computed P&L summary data (used for structure/metadata).
        """
        ws = wb.create_sheet("P&L Summary")
        years = pl_summary.years
        num_years = len(years)

        # Column references in All Transactions sheet:
        # F = Amount, D = Category, P = Category Type, Q = Year
        txn_sheet = f"'{self.SHEET_ALL_TRANSACTIONS}'"

        row = 1

        # Report metadata header
        ws.cell(row=row, column=1, value="REPORT SUMMARY")
        ws.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 1
        ws.cell(row=row, column=1, value="Period")
        ws.cell(row=row, column=2, value=pl_summary.period_display)
        row += 1
        ws.cell(row=row, column=1, value="Accounts")
        ws.cell(row=row, column=2, value=pl_summary.accounts_display)
        row += 2

        # Income section header with year columns
        ws.cell(row=row, column=1, value="INCOME")
        ws.cell(row=row, column=1).font = Font(bold=True)
        for i, year in enumerate(years):
            cell = ws.cell(row=row, column=i + 2, value=str(year))
            cell.font = Font(bold=True)
        cell = ws.cell(row=row, column=num_years + 2, value="Total")
        cell.font = Font(bold=True)
        income_header_row = row
        row += 1

        # Income categories with SUMIFS formulas
        income_start_row = row
        for cat in pl_summary.all_income_categories:
            ws.cell(row=row, column=1, value=sanitize_for_csv(cat))
            for i, year in enumerate(years):
                # SUMIFS: Amount where Category Type = "income", Category = cat, Year = year
                formula = (
                    f"=SUMIFS({txn_sheet}!$F:$F,"
                    f"{txn_sheet}!$P:$P,\"income\","
                    f"{txn_sheet}!$D:$D,A{row},"
                    f"{txn_sheet}!$Q:$Q,{year})"
                )
                cell = ws.cell(row=row, column=i + 2, value=formula)
                cell.number_format = self._money_format()
            # Total column: sum of year columns
            year_cols = [get_column_letter(i + 2) for i in range(num_years)]
            total_formula = f"=SUM({year_cols[0]}{row}:{year_cols[-1]}{row})" if year_cols else "=0"
            cell = ws.cell(row=row, column=num_years + 2, value=total_formula)
            cell.number_format = self._money_format()
            row += 1
        income_end_row = row - 1

        # Total Income row (sum of income category rows)
        ws.cell(row=row, column=1, value="Total Income")
        ws.cell(row=row, column=1).font = Font(bold=True)
        total_income_row = row
        for i in range(num_years + 1):
            col = i + 2
            col_letter = get_column_letter(col)
            if income_start_row <= income_end_row:
                formula = f"=SUM({col_letter}{income_start_row}:{col_letter}{income_end_row})"
            else:
                formula = "=0"
            cell = ws.cell(row=row, column=col, value=formula)
            cell.font = Font(bold=True)
            cell.number_format = self._money_format()
        row += 2

        # Expense section header with year columns
        ws.cell(row=row, column=1, value="EXPENSES")
        ws.cell(row=row, column=1).font = Font(bold=True)
        for i, year in enumerate(years):
            cell = ws.cell(row=row, column=i + 2, value=str(year))
            cell.font = Font(bold=True)
        cell = ws.cell(row=row, column=num_years + 2, value="Total")
        cell.font = Font(bold=True)
        row += 1

        # Expense categories with SUMIFS formulas (negate to show positive)
        expense_start_row = row
        for cat in pl_summary.all_expense_categories:
            ws.cell(row=row, column=1, value=sanitize_for_csv(cat))
            for i, year in enumerate(years):
                # Negate SUMIFS since expenses are stored as negative amounts
                formula = (
                    f"=-SUMIFS({txn_sheet}!$F:$F,"
                    f"{txn_sheet}!$P:$P,\"expense\","
                    f"{txn_sheet}!$D:$D,A{row},"
                    f"{txn_sheet}!$Q:$Q,{year})"
                )
                cell = ws.cell(row=row, column=i + 2, value=formula)
                cell.number_format = self._money_format()
            # Total column
            year_cols = [get_column_letter(i + 2) for i in range(num_years)]
            total_formula = f"=SUM({year_cols[0]}{row}:{year_cols[-1]}{row})" if year_cols else "=0"
            cell = ws.cell(row=row, column=num_years + 2, value=total_formula)
            cell.number_format = self._money_format()
            row += 1
        expense_end_row = row - 1

        # Total Expenses row
        ws.cell(row=row, column=1, value="Total Expenses")
        ws.cell(row=row, column=1).font = Font(bold=True)
        total_expense_row = row
        for i in range(num_years + 1):
            col = i + 2
            col_letter = get_column_letter(col)
            if expense_start_row <= expense_end_row:
                formula = f"=SUM({col_letter}{expense_start_row}:{col_letter}{expense_end_row})"
            else:
                formula = "=0"
            cell = ws.cell(row=row, column=col, value=formula)
            cell.font = Font(bold=True)
            cell.number_format = self._money_format()
        row += 2

        # Net income row (Total Income - Total Expenses)
        ws.cell(row=row, column=1, value="NET INCOME")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        for i in range(num_years + 1):
            col = i + 2
            col_letter = get_column_letter(col)
            formula = f"={col_letter}{total_income_row}-{col_letter}{total_expense_row}"
            cell = ws.cell(row=row, column=col, value=formula)
            cell.font = Font(bold=True, size=12)
            cell.number_format = self._money_format()
        row += 3

        # Transfers memo section
        ws.cell(row=row, column=1, value="TRANSFERS")
        ws.cell(row=row, column=1).font = Font(bold=True, italic=True)
        for i, year in enumerate(years):
            cell = ws.cell(row=row, column=i + 2, value=str(year))
            cell.font = Font(bold=True, italic=True)
        cell = ws.cell(row=row, column=num_years + 2, value="Total")
        cell.font = Font(bold=True, italic=True)
        row += 1
        ws.cell(row=row, column=1, value="(Money moved between accounts - not counted as income or expense)")
        ws.cell(row=row, column=1).font = Font(italic=True, color="666666")
        row += 1

        # Transfer categories with SUMIFS formulas
        transfer_start_row = row
        for cat in pl_summary.all_transfer_categories:
            ws.cell(row=row, column=1, value=sanitize_for_csv(cat))
            for i, year in enumerate(years):
                # Transfers: use ABS to show absolute value
                formula = (
                    f"=ABS(SUMIFS({txn_sheet}!$F:$F,"
                    f"{txn_sheet}!$P:$P,\"transfer\","
                    f"{txn_sheet}!$D:$D,A{row},"
                    f"{txn_sheet}!$Q:$Q,{year}))"
                )
                cell = ws.cell(row=row, column=i + 2, value=formula)
                cell.number_format = self._money_format()
            # Total column
            year_cols = [get_column_letter(i + 2) for i in range(num_years)]
            total_formula = f"=SUM({year_cols[0]}{row}:{year_cols[-1]}{row})" if year_cols else "=0"
            cell = ws.cell(row=row, column=num_years + 2, value=total_formula)
            cell.number_format = self._money_format()
            row += 1
        transfer_end_row = row - 1

        # Total Transfers row
        ws.cell(row=row, column=1, value="Total Transfers")
        for i in range(num_years + 1):
            col = i + 2
            col_letter = get_column_letter(col)
            if transfer_start_row <= transfer_end_row:
                formula = f"=SUM({col_letter}{transfer_start_row}:{col_letter}{transfer_end_row})"
            else:
                formula = "=0"
            cell = ws.cell(row=row, column=col, value=formula)
            cell.number_format = self._money_format()

        # Adjust column widths
        ws.column_dimensions["A"].width = 40
        for i in range(num_years + 1):
            ws.column_dimensions[get_column_letter(i + 2)].width = 15

    def _create_master_list(
        self, wb: "Workbook", transactions: list[Transaction]
    ) -> None:
        """Create Master List (All Transactions) sheet.

        Args:
            wb: Workbook to add sheet to.
            transactions: Transaction data.
        """
        ws = wb.create_sheet(self.SHEET_ALL_TRANSACTIONS)

        # Headers with confidence scoring columns, fingerprint, and formula columns
        headers = [
            "Date", "Account", "Description", "Category", "Sub-category",
            "Amount", "Balance", "Source File", "Duplicate Flag", "Uncategorized Flag",
            "Confidence", "Matched Pattern", "Category Source", "Confidence Factors",
            "Fingerprint", "Category Type", "Year", "Year-Month"
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.centered

        # Sort transactions by date, then account
        sorted_txns = sorted(
            transactions, key=self._txn_sort_key
        )

        # Conditional formatting colors for confidence
        low_conf_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Red
        med_conf_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # Yellow
        high_conf_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # Green

        # Write data
        for row, txn in enumerate(sorted_txns, 2):
            ws.cell(row=row, column=1, value=txn.date)
            ws.cell(row=row, column=2, value=sanitize_for_csv(txn.account_name))
            ws.cell(row=row, column=3, value=sanitize_for_csv(txn.description))
            ws.cell(row=row, column=4, value=sanitize_for_csv(self._get_category_name(txn.category)))
            ws.cell(row=row, column=5, value=sanitize_for_csv(self._get_category_name(txn.subcategory)))

            amount_cell = ws.cell(row=row, column=6, value=float(txn.amount))
            amount_cell.number_format = self._money_format()
            if txn.amount < 0:
                amount_cell.font = self.money_negative
            else:
                amount_cell.font = self.money_positive

            if txn.running_balance is not None:
                balance_cell = ws.cell(row=row, column=7, value=float(txn.running_balance))
                balance_cell.number_format = self._money_format()

            ws.cell(row=row, column=8, value=sanitize_for_csv(txn.source_file))
            ws.cell(row=row, column=9, value="Yes" if txn.is_duplicate else "")
            ws.cell(row=row, column=10, value="Yes" if txn.is_uncategorized else "")

            # Confidence scoring columns
            if not txn.is_uncategorized:
                conf_cell = ws.cell(row=row, column=11, value=txn.confidence_score)
                conf_cell.number_format = "0.00"

                # Apply conditional formatting based on confidence
                if txn.confidence_score < 0.6:
                    conf_cell.fill = low_conf_fill
                elif txn.confidence_score < 0.8:
                    conf_cell.fill = med_conf_fill
                else:
                    conf_cell.fill = high_conf_fill

            ws.cell(row=row, column=12, value=sanitize_for_csv(txn.matched_pattern or ""))
            ws.cell(row=row, column=13, value=sanitize_for_csv(txn.category_source))

            # Format confidence factors as semicolon-separated list
            factors_str = "; ".join(txn.confidence_factors) if txn.confidence_factors else ""
            ws.cell(row=row, column=14, value=sanitize_for_csv(factors_str))

            # Fingerprint for correction matching
            ws.cell(row=row, column=15, value=txn.fingerprint)

            # Category Type formula (VLOOKUP from Category Lookup sheet)
            # D column = Category, lookup returns type from column B of Category Lookup
            ws.cell(row=row, column=16, value=f"=IFERROR(VLOOKUP(D{row},'Category Lookup'!$A:$B,2,FALSE),\"\")")

            # Year formula from Date column
            ws.cell(row=row, column=17, value=f"=YEAR(A{row})")

            # Year-Month formula for Category Analysis (e.g., "2023-01")
            ws.cell(row=row, column=18, value=f"=TEXT(A{row},\"YYYY-MM\")")

        # Adjust column widths
        widths = [12, 25, 40, 20, 20, 12, 12, 25, 12, 15, 10, 20, 12, 40, 18, 14, 8, 10]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Add data validation dropdown on Category column (D)
        # Reference category names from Category Lookup sheet
        num_categories = len([c for c in self.config.categories.values() if not c.parent_id])
        if num_categories > 0 and len(sorted_txns) > 0:
            dv = DataValidation(
                type="list",
                formula1=f"'Category Lookup'!$A$2:$A${num_categories + 1}",
                allow_blank=True,
                showDropDown=False,  # False means show the dropdown arrow
            )
            dv.error = "Please select a valid category from the list"
            dv.errorTitle = "Invalid Category"
            dv.prompt = "Select a category"
            dv.promptTitle = "Category"
            # Apply to all data rows in the Category column
            dv.add(f"D2:D{len(sorted_txns) + 1}")
            ws.add_data_validation(dv)

        # Freeze header row
        ws.freeze_panes = "A2"

    def _create_review_queue(
        self, wb: "Workbook", transactions: list[Transaction]
    ) -> None:
        """Create Review Queue sheet sorted by confidence score (lowest first).

        This sheet helps users prioritize manual review of transactions by showing
        those with the lowest confidence scores at the top. Includes a reference
        to the row number in All Transactions for easy lookup.

        Args:
            wb: Workbook to add sheet to.
            transactions: Transaction data.
        """
        ws = wb.create_sheet("Review Queue")

        # Conditional formatting colors for confidence
        low_conf_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        med_conf_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        high_conf_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

        # Headers
        headers = [
            "Txn Row#", "Date", "Account", "Description", "Category",
            "Amount", "Confidence", "Matched Pattern", "Category Source",
            "Confidence Factors", "Fingerprint"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.centered

        # Sort transactions by confidence score (lowest first), then by date
        # We need to track the original row numbers in All Transactions
        # All Transactions is sorted by: date, account_name, description
        sorted_for_master = sorted(
            transactions, key=self._txn_sort_key
        )

        # Create a mapping of transaction ID to row number in All Transactions.
        # Use txn.id (UUID) for row mapping because identical transactions intentionally
        # share fingerprints (by design for correction matching), but each needs its own
        # row reference in the spreadsheet.
        txn_row_map = {txn.id: idx + 2 for idx, txn in enumerate(sorted_for_master)}

        # Sort by confidence score (lowest first) for review queue
        sorted_for_review = sorted(
            transactions, key=lambda t: (t.confidence_score, t.date, t.account_name)
        )

        # Write data
        for row, txn in enumerate(sorted_for_review, 2):
            # Row number in All Transactions (for reference)
            all_txn_row = txn_row_map.get(txn.id, "")
            ws.cell(row=row, column=1, value=all_txn_row)

            ws.cell(row=row, column=2, value=txn.date)
            ws.cell(row=row, column=3, value=sanitize_for_csv(txn.account_name))
            ws.cell(row=row, column=4, value=sanitize_for_csv(txn.description))
            ws.cell(row=row, column=5, value=sanitize_for_csv(self._get_category_name(txn.category)))

            amount_cell = ws.cell(row=row, column=6, value=float(txn.amount))
            amount_cell.number_format = self._money_format()
            if txn.amount < 0:
                amount_cell.font = self.money_negative
            else:
                amount_cell.font = self.money_positive

            # Confidence score with conditional formatting
            conf_cell = ws.cell(row=row, column=7, value=txn.confidence_score)
            conf_cell.number_format = "0.00"
            if txn.confidence_score < 0.6:
                conf_cell.fill = low_conf_fill
            elif txn.confidence_score < 0.8:
                conf_cell.fill = med_conf_fill
            else:
                conf_cell.fill = high_conf_fill

            ws.cell(row=row, column=8, value=sanitize_for_csv(txn.matched_pattern or ""))
            ws.cell(row=row, column=9, value=sanitize_for_csv(txn.category_source))

            factors_str = "; ".join(txn.confidence_factors) if txn.confidence_factors else ""
            ws.cell(row=row, column=10, value=sanitize_for_csv(factors_str))

            ws.cell(row=row, column=11, value=txn.fingerprint)

        # Adjust column widths
        widths = [10, 12, 20, 40, 20, 12, 10, 20, 12, 40, 18]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        logger.debug(f"Created Review Queue with {len(transactions)} transactions")

    def _create_deposits_sheet(
        self, wb: "Workbook", transactions: list[Transaction]
    ) -> None:
        """Create Deposits sheet showing only positive amount transactions.

        Args:
            wb: Workbook to add sheet to.
            transactions: Transaction data.
        """
        ws = wb.create_sheet("Deposits")

        # Headers
        headers = [
            "Date", "Account", "Description", "Category", "Sub-category",
            "Amount", "Balance", "Source File"
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.centered

        # Filter to deposits only (amount > 0, excludes zero and negative)
        deposits = [txn for txn in transactions if txn.amount > 0]

        # Sort deposits by date, then account, then description
        sorted_deposits = sorted(
            deposits, key=self._txn_sort_key
        )

        # Write data
        for row, txn in enumerate(sorted_deposits, 2):
            ws.cell(row=row, column=1, value=txn.date)
            ws.cell(row=row, column=2, value=sanitize_for_csv(txn.account_name))
            ws.cell(row=row, column=3, value=sanitize_for_csv(txn.description))
            ws.cell(row=row, column=4, value=sanitize_for_csv(self._get_category_name(txn.category)))
            ws.cell(row=row, column=5, value=sanitize_for_csv(self._get_category_name(txn.subcategory)))

            amount_cell = ws.cell(row=row, column=6, value=float(txn.amount))
            amount_cell.number_format = self._money_format()
            amount_cell.font = self.money_positive  # Always green since deposits are positive

            if txn.running_balance is not None:
                balance_cell = ws.cell(row=row, column=7, value=float(txn.running_balance))
                balance_cell.number_format = self._money_format()

            ws.cell(row=row, column=8, value=sanitize_for_csv(txn.source_file))

        # Adjust column widths
        widths = [12, 25, 40, 20, 20, 12, 12, 25]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

    def _create_transfers_sheet(
        self, wb: "Workbook", transactions: list[Transaction]
    ) -> None:
        """Create Transfers sheet showing only transfer-type transactions.

        Args:
            wb: Workbook to add sheet to.
            transactions: Transaction data.
        """
        ws = wb.create_sheet("Transfers")

        # Headers
        headers = [
            "Date", "Account", "Description", "Category", "Sub-category",
            "Amount", "Balance", "Source File"
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.centered

        # Filter to transfers only (category type == "transfer")
        transfers = [
            txn for txn in transactions
            if self._get_category_type(txn.category) == "transfer"
        ]

        # Sort transfers by date, then account, then description
        sorted_transfers = sorted(
            transfers, key=self._txn_sort_key
        )

        # Write data
        for row, txn in enumerate(sorted_transfers, 2):
            ws.cell(row=row, column=1, value=txn.date)
            ws.cell(row=row, column=2, value=sanitize_for_csv(txn.account_name))
            ws.cell(row=row, column=3, value=sanitize_for_csv(txn.description))
            ws.cell(row=row, column=4, value=sanitize_for_csv(self._get_category_name(txn.category)))
            ws.cell(row=row, column=5, value=sanitize_for_csv(self._get_category_name(txn.subcategory)))

            amount_cell = ws.cell(row=row, column=6, value=float(txn.amount))
            amount_cell.number_format = self._money_format()
            # Green for positive, red for negative
            amount_cell.font = self.money_positive if txn.amount >= 0 else self.money_negative

            if txn.running_balance is not None:
                balance_cell = ws.cell(row=row, column=7, value=float(txn.running_balance))
                balance_cell.number_format = self._money_format()

            ws.cell(row=row, column=8, value=sanitize_for_csv(txn.source_file))

        # Adjust column widths
        widths = [12, 25, 40, 20, 20, 12, 12, 25]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

    def _create_account_sheets(
        self, wb: "Workbook", transactions: list[Transaction]
    ) -> None:
        """Create per-account transaction sheets.

        Args:
            wb: Workbook to add sheets to.
            transactions: Transaction data.
        """
        # Group by account
        by_account: dict[str, list[Transaction]] = {}
        for txn in transactions:
            if txn.account_name not in by_account:
                by_account[txn.account_name] = []
            by_account[txn.account_name].append(txn)

        for account_name, account_txns in sorted(by_account.items()):
            # Sanitize and truncate sheet name (Excel limit is 31 chars)
            # Excel doesn't allow: * ? / \ [ ] in sheet names
            sheet_name = account_name
            for char in ['*', '?', '/', '\\', '[', ']', ':']:
                sheet_name = sheet_name.replace(char, '')
            sheet_name = sheet_name[:31]
            ws = wb.create_sheet(sheet_name)

            headers = ["Date", "Description", "Category", "Amount", "Balance"]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill

            # Sort by date
            sorted_txns = sorted(account_txns, key=lambda t: (t.date, t.description))

            for row, txn in enumerate(sorted_txns, 2):
                ws.cell(row=row, column=1, value=txn.date)
                ws.cell(row=row, column=2, value=sanitize_for_csv(txn.description))
                ws.cell(row=row, column=3, value=sanitize_for_csv(self._get_category_name(txn.category)))

                amount_cell = ws.cell(row=row, column=4, value=float(txn.amount))
                amount_cell.number_format = self._money_format()
                if txn.amount < 0:
                    amount_cell.font = self.money_negative
                else:
                    amount_cell.font = self.money_positive

                if txn.running_balance is not None:
                    balance_cell = ws.cell(row=row, column=5, value=float(txn.running_balance))
                    balance_cell.number_format = self._money_format()

            # Adjust widths
            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 40
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 12
            ws.column_dimensions["E"].width = 12

            ws.freeze_panes = "A2"

    def _create_account_summary(
        self, wb: "Workbook", transactions: list[Transaction]
    ) -> None:
        """Create Account Summary sheet showing balance overview per account.

        Args:
            wb: Workbook to add sheet to.
            transactions: Transaction data.
        """
        from collections import defaultdict

        ws = wb.create_sheet("Account Summary")

        # Headers using existing instance variables
        headers = ["Account", "Opening Balance", "Total Credits", "Total Debits", "Closing Balance"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.centered

        # Group transactions by account_id
        by_account: dict[str, list[Transaction]] = defaultdict(list)
        for txn in transactions:
            by_account[txn.account_id].append(txn)

        # Warn about orphaned transactions (account_id not in config)
        config_account_ids = set(self.config.accounts.keys())
        orphaned_account_ids = set(by_account.keys()) - config_account_ids
        if orphaned_account_ids:
            orphaned_txn_count = sum(len(by_account[aid]) for aid in orphaned_account_ids)
            logger.warning(
                f"Account Summary: {orphaned_txn_count} transaction(s) from "
                f"{len(orphaned_account_ids)} unknown account(s) not included: "
                f"{orphaned_account_ids}. "
                f"Add these accounts to accounts.yaml to include them in the summary."
            )

        # Sort accounts by display_order, then name
        sorted_accounts = sorted(
            self.config.accounts.items(),
            key=lambda x: (x[1].display_order, x[1].name)
        )

        # Calculate per account
        # Note: opening_balance may be None if not explicitly set and balance inference
        # didn't run or couldn't infer a value. We default to 0, which produces
        # mathematically consistent closing balances (Opening + Credits + Debits) but may
        # not reflect actual account balances unless opening_balance is truly zero or
        # explicitly set. Users should set opening_balance for accounts with pre-existing
        # balances (credit cards, loans, or checking accounts opened before tracking began).
        for account_id, account in sorted_accounts:
            txns = by_account.get(account_id, [])
            opening = account.opening_balance if account.opening_balance is not None else Decimal("0")
            credits = sum(t.amount for t in txns if t.amount > 0)
            debits = sum(t.amount for t in txns if t.amount < 0)
            closing = opening + credits + debits

            ws.append([
                account.name,
                float(opening),
                float(credits),
                float(debits),
                float(closing),
            ])

        # Format currency columns (B through E)
        money_fmt = self._money_format()
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=5):
            for cell in row:
                cell.number_format = money_fmt
                cell.alignment = Alignment(horizontal="right")

        # Set column widths
        ws.column_dimensions["A"].width = 25
        for col_letter in ["B", "C", "D", "E"]:
            ws.column_dimensions[col_letter].width = 18

        # Add totals row (only if there are data rows to sum)
        if ws.max_row > 1:
            last_row = ws.max_row + 1
            ws.cell(row=last_row, column=1, value="TOTAL").font = Font(bold=True)

            for col in range(2, 6):
                col_letter = get_column_letter(col)
                formula = f"=SUM({col_letter}2:{col_letter}{last_row - 1})"
                cell = ws.cell(row=last_row, column=col, value=formula)
                cell.number_format = money_fmt
                cell.font = Font(bold=True)

        # Freeze header row
        ws.freeze_panes = "A2"

    def _create_category_analysis(
        self, wb: "Workbook", transactions: list[Transaction]
    ) -> None:
        """Create Category Analysis sheet with SUMIFS formulas.

        Uses formulas that reference 'All Transactions' sheet so changes
        to category assignments automatically update monthly totals.

        Args:
            wb: Workbook to add sheet to.
            transactions: Transaction data.
        """
        ws = wb.create_sheet("Category Analysis")

        # Column references in All Transactions sheet:
        # F = Amount, D = Category, R = Year-Month
        txn_sheet = f"'{self.SHEET_ALL_TRANSACTIONS}'"

        # Get unique categories and months from transactions
        categories: set[str] = set()
        months: set[str] = set()

        for txn in transactions:
            cat_name = self._get_category_name(txn.category)
            if cat_name:
                categories.add(cat_name)
            else:
                categories.add("Uncategorized")
            months.add(txn.date.strftime("%Y-%m"))

        all_categories = sorted(categories)
        all_months = sorted(months)

        if not all_months or not all_categories:
            ws.cell(row=1, column=1, value="No transaction data")
            return

        # Write headers
        ws.cell(row=1, column=1, value="Category")
        ws.cell(row=1, column=1).font = self.header_font
        ws.cell(row=1, column=1).fill = self.header_fill

        for col, month in enumerate(all_months, 2):
            cell = ws.cell(row=1, column=col, value=month)
            cell.font = self.header_font
            cell.fill = self.header_fill

        # Total column
        total_col = len(all_months) + 2
        ws.cell(row=1, column=total_col, value="Total")
        ws.cell(row=1, column=total_col).font = self.header_font
        ws.cell(row=1, column=total_col).fill = self.header_fill

        # Write data with SUMIFS formulas
        for row, cat_name in enumerate(all_categories, 2):
            ws.cell(row=row, column=1, value=sanitize_for_csv(cat_name))

            for col, month in enumerate(all_months, 2):
                # SUMIFS: Amount where Category = cat and Year-Month = month
                formula = (
                    f"=SUMIFS({txn_sheet}!$F:$F,"
                    f"{txn_sheet}!$D:$D,A{row},"
                    f"{txn_sheet}!$R:$R,\"{month}\")"
                )
                cell = ws.cell(row=row, column=col, value=formula)
                cell.number_format = self._money_format()

            # Total column: sum of month columns for this row
            first_col = get_column_letter(2)
            last_col = get_column_letter(len(all_months) + 1)
            total_formula = f"=SUM({first_col}{row}:{last_col}{row})"
            cell = ws.cell(row=row, column=total_col, value=total_formula)
            cell.number_format = self._money_format()
            cell.font = Font(bold=True)

        # Adjust widths
        ws.column_dimensions["A"].width = 25
        for i in range(len(all_months) + 1):
            ws.column_dimensions[get_column_letter(i + 2)].width = 12

        ws.freeze_panes = "B2"

    def _create_anomalies_sheet(
        self,
        wb: "Workbook",
        transactions: list[Transaction],
        date_gaps: list[dict[str, object]],
    ) -> None:
        """Create Anomalies sheet.

        Args:
            wb: Workbook to add sheet to.
            transactions: Transaction data.
            date_gaps: Date gap anomalies.
        """
        ws = wb.create_sheet("Anomalies")

        # Transaction anomalies
        ws.cell(row=1, column=1, value="Transaction Anomalies")
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)

        headers = ["Date", "Account", "Description", "Amount", "Reason"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill

        row = 3
        anomaly_txns = [t for t in transactions if t.is_anomaly]
        for txn in sorted(anomaly_txns, key=lambda t: t.date):
            ws.cell(row=row, column=1, value=txn.date)
            ws.cell(row=row, column=2, value=sanitize_for_csv(txn.account_name))
            ws.cell(row=row, column=3, value=sanitize_for_csv(txn.description))

            amount_cell = ws.cell(row=row, column=4, value=float(txn.amount))
            amount_cell.number_format = self._money_format()

            ws.cell(row=row, column=5, value=sanitize_for_csv("; ".join(txn.anomaly_reasons)))
            row += 1

        # Date gap anomalies
        row += 2
        ws.cell(row=row, column=1, value="Date Gap Anomalies")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        gap_headers = ["Account", "Start Date", "End Date", "Gap (Days)", "Severity"]
        for col, header in enumerate(gap_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
        row += 1

        for gap in date_gaps:
            ws.cell(row=row, column=1, value=str(gap.get("account_id", "")))
            ws.cell(row=row, column=2, value=gap.get("start_date"))
            ws.cell(row=row, column=3, value=gap.get("end_date"))
            ws.cell(row=row, column=4, value=gap.get("gap_days"))
            ws.cell(row=row, column=5, value=str(gap.get("severity", "")))
            row += 1

        # Adjust widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 40

    def _get_category_type(self, category_id: str | None) -> str | None:
        """Get the type (income/expense/transfer) for a category.

        Args:
            category_id: Category ID.

        Returns:
            Category type or None.
        """
        if not category_id:
            return None

        category = self.config.categories.get(category_id)
        if category:
            return category.category_type.value if category.category_type else None

        return None

    def _get_category_name(self, category_id: str | None) -> str | None:
        """Get display name for a category.

        Args:
            category_id: Category ID.

        Returns:
            Category name or None.
        """
        if not category_id:
            return None

        category = self.config.categories.get(category_id)
        if category:
            return category.name

        return category_id

    def _money_format(self) -> str:
        """Get number format for money values.

        Returns:
            Excel number format string.
        """
        symbol = self.output_config.currency_symbol
        return f'{symbol}#,##0.00_);[Red]({symbol}#,##0.00)'
