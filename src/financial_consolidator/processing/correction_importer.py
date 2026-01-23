"""Import category corrections from reviewed output files."""

import csv
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

from financial_consolidator.config import Config
from financial_consolidator.models.category import CategoryCorrection
from financial_consolidator.utils.logging_config import get_logger

logger = get_logger(__name__)


class CorrectionImportError(Exception):
    """Exception raised when correction import fails."""

    def __init__(self, message: str, file_path: Path | None = None):
        """Initialize CorrectionImportError.

        Args:
            message: Error message.
            file_path: Optional path to the file that failed to import.
        """
        self.file_path = file_path
        super().__init__(message)


@dataclass
class ImportResult:
    """Result of a correction import operation.

    Attributes:
        imported_count: Number of corrections successfully imported.
        skipped_count: Number of corrections skipped.
        skipped_reasons: List of reasons for skipped corrections.
        corrections: Dictionary of fingerprint to CategoryCorrection.
    """

    imported_count: int
    skipped_count: int
    skipped_reasons: list[str]
    corrections: dict[str, CategoryCorrection]


class CorrectionImporter:
    """Imports category corrections from reviewed XLSX or CSV output files.

    The importer reads the output file, finds transactions where the category
    has been manually changed, and creates corrections that can be applied
    to future analysis runs.

    Supported file formats:
    - CSV: all_transactions.csv
    - XLSX: All Transactions sheet from analysis output

    Required columns:
    - Fingerprint: Transaction identifier for matching
    - Category: The corrected category name
    """

    def __init__(self, config: Config):
        """Initialize the CorrectionImporter.

        Args:
            config: Application configuration with category definitions.
        """
        self.config = config

    def import_corrections(self, file_path: Path) -> ImportResult:
        """Import corrections from a file.

        Args:
            file_path: Path to the XLSX or CSV file.

        Returns:
            ImportResult with statistics and corrections.

        Raises:
            CorrectionImportError: If the file cannot be read or parsed.
            FileNotFoundError: If the file doesn't exist.
        """
        if not file_path.exists():
            raise FileNotFoundError(f"Corrections file not found: {file_path}")

        suffix = file_path.suffix.lower()

        if suffix == ".csv":
            return self._import_from_csv(file_path)
        elif suffix in (".xlsx", ".xls"):
            return self._import_from_xlsx(file_path)
        else:
            raise CorrectionImportError(
                f"Unsupported file format: {suffix}. Use .csv or .xlsx",
                file_path
            )

    def _import_from_csv(self, file_path: Path) -> ImportResult:
        """Import corrections from a CSV file.

        Args:
            file_path: Path to the CSV file.

        Returns:
            ImportResult with statistics and corrections.
        """
        corrections: dict[str, CategoryCorrection] = {}
        skipped_reasons: list[str] = []
        imported_count = 0
        skipped_count = 0

        try:
            # Use utf-8-sig encoding to handle BOM (Byte Order Mark)
            with open(file_path, encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)

                # Check for required columns
                if not reader.fieldnames:
                    raise CorrectionImportError(
                        "CSV file has no headers",
                        file_path
                    )

                # Handle None values in fieldnames (can occur with malformed CSV)
                headers = [h.lower() if h else "" for h in reader.fieldnames]
                if "fingerprint" not in headers:
                    raise CorrectionImportError(
                        "CSV file missing required 'Fingerprint' column",
                        file_path
                    )
                if "category" not in headers:
                    raise CorrectionImportError(
                        "CSV file missing required 'Category' column",
                        file_path
                    )

                # Find actual column names (case-insensitive, handle None)
                fingerprint_col = next(
                    h for h in reader.fieldnames if h and h.lower() == "fingerprint"
                )
                category_col = next(
                    h for h in reader.fieldnames if h and h.lower() == "category"
                )
                subcategory_col = next(
                    (h for h in reader.fieldnames if h and h.lower() == "sub-category"),
                    None
                )
                source_col = next(
                    (h for h in reader.fieldnames if h and h.lower() == "category source"),
                    None
                )

                for row_num, row in enumerate(reader, start=2):
                    result = self._process_row(
                        row, fingerprint_col, category_col, subcategory_col,
                        source_col, row_num, file_path
                    )

                    if result is None:
                        skipped_count += 1
                        skipped_reasons.append(f"Row {row_num}: Empty fingerprint or category")
                    elif isinstance(result, str):
                        skipped_count += 1
                        skipped_reasons.append(result)
                    else:
                        corrections[result.fingerprint] = result
                        imported_count += 1

        except csv.Error as e:
            raise CorrectionImportError(f"CSV parse error: {e}", file_path)
        except UnicodeDecodeError as e:
            raise CorrectionImportError(f"File encoding error: {e}", file_path)
        except OSError as e:
            raise CorrectionImportError(f"Failed to read CSV file: {e}", file_path)

        logger.info(
            f"Imported {imported_count} corrections from CSV, "
            f"skipped {skipped_count}"
        )

        return ImportResult(
            imported_count=imported_count,
            skipped_count=skipped_count,
            skipped_reasons=skipped_reasons,
            corrections=corrections
        )

    def _import_from_xlsx(self, file_path: Path) -> ImportResult:
        """Import corrections from an XLSX file.

        Args:
            file_path: Path to the XLSX file.

        Returns:
            ImportResult with statistics and corrections.
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise CorrectionImportError(
                "openpyxl library not installed. Install with: pip install openpyxl",
                file_path
            )

        corrections: dict[str, CategoryCorrection] = {}
        skipped_reasons: list[str] = []
        imported_count = 0
        skipped_count = 0

        wb = None
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)

            # Find All Transactions sheet
            sheet_name = None
            for name in wb.sheetnames:
                if name.lower() == "all transactions":
                    sheet_name = name
                    break

            if not sheet_name:
                raise CorrectionImportError(
                    "XLSX file missing 'All Transactions' sheet",
                    file_path
                )

            ws = wb[sheet_name]

            # Read header row to find column indices
            headers: list[str] = []
            for cell in ws[1]:
                headers.append(str(cell.value or "").lower())

            if "fingerprint" not in headers:
                raise CorrectionImportError(
                    "Sheet missing required 'Fingerprint' column",
                    file_path
                )
            if "category" not in headers:
                raise CorrectionImportError(
                    "Sheet missing required 'Category' column",
                    file_path
                )

            fingerprint_idx = headers.index("fingerprint")
            category_idx = headers.index("category")
            subcategory_idx = headers.index("sub-category") if "sub-category" in headers else None
            source_idx = headers.index("category source") if "category source" in headers else None

            # Process data rows
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or len(row) <= max(fingerprint_idx, category_idx):
                    continue

                fingerprint = str(row[fingerprint_idx] or "").strip()
                category = str(row[category_idx] or "").strip()
                subcategory = str(row[subcategory_idx] or "").strip() if subcategory_idx is not None and len(row) > subcategory_idx else ""
                source = str(row[source_idx] or "").strip() if source_idx is not None and len(row) > source_idx else ""

                row_dict = {
                    "fingerprint": fingerprint,
                    "category": category,
                    "sub-category": subcategory,
                    "category source": source
                }

                result = self._process_row(
                    row_dict, "fingerprint", "category", "sub-category",
                    "category source", row_num, file_path
                )

                if result is None:
                    skipped_count += 1
                    skipped_reasons.append(f"Row {row_num}: Empty fingerprint or category")
                elif isinstance(result, str):
                    skipped_count += 1
                    skipped_reasons.append(result)
                else:
                    corrections[result.fingerprint] = result
                    imported_count += 1

        except CorrectionImportError:
            raise
        except (OSError, PermissionError) as e:
            raise CorrectionImportError(f"Failed to read XLSX file: {e}", file_path)
        except (KeyError, ValueError, TypeError) as e:
            raise CorrectionImportError(f"Invalid XLSX format: {e}", file_path)
        finally:
            if wb is not None:
                wb.close()

        logger.info(
            f"Imported {imported_count} corrections from XLSX, "
            f"skipped {skipped_count}"
        )

        return ImportResult(
            imported_count=imported_count,
            skipped_count=skipped_count,
            skipped_reasons=skipped_reasons,
            corrections=corrections
        )

    def _process_row(
        self,
        row: dict[str, object],
        fingerprint_col: str,
        category_col: str,
        subcategory_col: str | None,
        source_col: str | None,
        row_num: int,
        file_path: Path
    ) -> CategoryCorrection | str | None:
        """Process a single row and return a correction or skip reason.

        Args:
            row: Row data dictionary.
            fingerprint_col: Column name for fingerprint.
            category_col: Column name for category.
            subcategory_col: Optional column name for subcategory.
            source_col: Optional column name for category source.
            row_num: Row number for error messages.
            file_path: File path for error messages.

        Returns:
            CategoryCorrection if valid, string error message if invalid,
            None if row should be skipped silently.
        """
        fingerprint = str(row.get(fingerprint_col, "") or "").strip()
        category_name = str(row.get(category_col, "") or "").strip()

        # Skip empty rows
        if not fingerprint or not category_name:
            return None

        # Validate fingerprint format (should be 16-character hex string)
        if not re.match(r"^[0-9a-fA-F]{16}$", fingerprint):
            # Show full value for debugging (fingerprints are 16 chars, limit display for malformed data)
            display_fp = fingerprint if len(fingerprint) <= 32 else f"{fingerprint[:32]}..."
            return f"Row {row_num}: Invalid fingerprint format '{display_fp}'"

        # Normalize to lowercase for consistent dictionary lookups
        # (Transaction.fingerprint generates lowercase hex)
        fingerprint = fingerprint.lower()

        # Map category name to ID
        category_id = self.config.get_category_id_by_name(category_name)
        if not category_id:
            return f"Row {row_num}: Unknown category '{category_name}'"

        # Get original source to check if it was already a correction
        original_source = ""
        if source_col:
            original_source = str(row.get(source_col, "") or "").strip()

        # Handle subcategory
        subcategory_id = None
        if subcategory_col:
            subcategory_name = str(row.get(subcategory_col, "") or "").strip()
            if subcategory_name:
                subcategory_id = self.config.get_category_id_by_name(subcategory_name)
                if not subcategory_id:
                    logger.warning(
                        f"Row {row_num}: Unknown subcategory '{subcategory_name}', ignoring"
                    )

        return CategoryCorrection(
            fingerprint=fingerprint,
            category_id=category_id,
            subcategory_id=subcategory_id,
            original_category_id=None,  # Not tracked in output
            original_source=original_source,
            corrected_at=datetime.now(timezone.utc).isoformat(),
            source_file=str(file_path)
        )


def import_corrections_from_file(
    file_path: Path,
    config: Config,
) -> ImportResult:
    """Convenience function to import corrections from a file.

    Args:
        file_path: Path to the XLSX or CSV file.
        config: Application configuration.

    Returns:
        ImportResult with statistics and corrections.
    """
    importer = CorrectionImporter(config)
    return importer.import_corrections(file_path)
