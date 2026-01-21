"""Excel parser using openpyxl library."""

from collections.abc import Sequence
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from financial_consolidator.models.transaction import RawTransaction, TransactionType
from financial_consolidator.parsers.base import BaseParser, ParseError
from financial_consolidator.utils.date_utils import parse_date
from financial_consolidator.utils.decimal_utils import parse_amount
from financial_consolidator.utils.logging_config import get_logger

logger = get_logger(__name__)

# Import openpyxl - handle import error gracefully
try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    load_workbook = None  # type: ignore
    Worksheet = None  # type: ignore

# Maximum Excel file size to prevent memory exhaustion (50 MB)
MAX_EXCEL_FILE_SIZE = 50 * 1024 * 1024


class ExcelParser(BaseParser):
    """Parser for Excel financial statement files.

    Supports .xlsx and .xls files. Uses similar column detection
    logic as the CSV parser.
    """

    @property
    def supported_extensions(self) -> list[str]:
        """Return supported file extensions."""
        return [".xlsx", ".xls"]

    def can_parse(self, file_path: Path) -> bool:
        """Check if this parser can handle the file.

        Args:
            file_path: Path to the file.

        Returns:
            True if file is a parseable Excel file.
        """
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl library not installed, Excel parsing unavailable")
            return False

        if not self._check_extension(file_path):
            return False

        # Try to open the file to verify it's a valid Excel file
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            wb.close()
            return True
        except Exception:
            return False

    def parse(self, file_path: Path) -> list[RawTransaction]:
        """Parse an Excel file and return raw transactions.

        Args:
            file_path: Path to the Excel file.

        Returns:
            List of RawTransaction objects.

        Raises:
            ParseError: If parsing fails.
        """
        if not OPENPYXL_AVAILABLE:
            raise ParseError("openpyxl library not installed", file_path)

        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        # Check file size to prevent memory exhaustion
        file_size = file_path.stat().st_size
        if file_size > MAX_EXCEL_FILE_SIZE:
            raise ParseError(
                f"File too large ({file_size / 1024 / 1024:.1f} MB). "
                f"Maximum allowed is {MAX_EXCEL_FILE_SIZE / 1024 / 1024:.0f} MB",
                file_path,
            )

        logger.info(f"Parsing Excel file: {file_path.name}")

        transactions = []
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)

            # Process each worksheet
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_txns = self._parse_worksheet(sheet, file_path.name, sheet_name)
                transactions.extend(sheet_txns)

            wb.close()

        except Exception as e:
            raise ParseError(f"Failed to parse Excel file: {e}", file_path) from e

        logger.info(f"Parsed {len(transactions)} transactions from {file_path.name}")
        return transactions

    def detect_institution(self, file_path: Path) -> str | None:
        """Detect financial institution from Excel content.

        Args:
            file_path: Path to the file.

        Returns:
            Institution name if detected.
        """
        # Could look at sheet names or specific cells for institution info
        # For now, return None
        return None

    def _parse_worksheet(
        self, sheet: "Worksheet", source_file: str, sheet_name: str
    ) -> list[RawTransaction]:
        """Parse a single worksheet.

        Args:
            sheet: Excel worksheet.
            source_file: Source file name.
            sheet_name: Name of the worksheet.

        Returns:
            List of RawTransaction objects.
        """
        transactions: list[RawTransaction] = []

        # Read all rows
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return transactions

        # Find header row
        header_row_idx = self._find_header_row(rows)
        if header_row_idx is None:
            logger.warning(f"No header row found in sheet '{sheet_name}'")
            return transactions

        headers = [
            str(h).lower().strip() if h else "" for h in rows[header_row_idx]
        ]

        # Detect column mapping
        mapping = self._detect_column_mapping(headers)
        if mapping is None:
            logger.warning(f"Could not detect column mapping in sheet '{sheet_name}'")
            return transactions

        # Parse data rows
        for row_num, row in enumerate(rows[header_row_idx + 1 :], start=1):
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            try:
                txn = self._parse_row(row, mapping, source_file, sheet_name)
                if txn:
                    transactions.append(txn)
            except Exception as e:
                logger.warning(
                    f"Error parsing row {row_num} in sheet '{sheet_name}': {e}"
                )

        return transactions

    def _find_header_row(
        self, rows: Sequence[tuple[object, ...]]
    ) -> int | None:
        """Find the header row in the worksheet.

        Args:
            rows: All rows from the worksheet.

        Returns:
            Index of header row or None.
        """
        header_keywords = [
            "date", "description", "amount", "debit", "credit", "balance",
            "transaction", "posted", "memo", "check", "type", "category",
        ]

        for i, row in enumerate(rows[:20]):  # Check first 20 rows
            if not row:
                continue

            # Count keyword matches
            text_cells = [
                str(cell).lower() for cell in row if cell is not None
            ]
            keyword_matches = sum(
                1 for cell in text_cells
                if any(kw in cell for kw in header_keywords)
            )

            # If at least 2 keywords match, this is likely the header
            if keyword_matches >= 2:
                return i

        return None

    def _detect_column_mapping(
        self, headers: list[str]
    ) -> dict[str, int] | None:
        """Detect column mapping from headers.

        Args:
            headers: List of header strings (lowercase).

        Returns:
            Mapping of field name to column index.
        """
        mapping: dict[str, int] = {}

        for i, header in enumerate(headers):
            if not header:
                continue

            # Date column
            if "date" not in mapping and any(
                kw in header for kw in ["date", "posted", "trans"]
            ):
                if "description" not in header:
                    mapping["date"] = i

            # Description column
            if "description" not in mapping and any(
                kw in header
                for kw in ["description", "desc", "memo", "payee", "merchant", "name"]
            ):
                mapping["description"] = i

            # Amount column (single)
            if "amount" not in mapping and "amount" in header:
                mapping["amount"] = i

            # Debit column
            if "debit" not in mapping and any(
                kw in header for kw in ["debit", "withdrawal", "payment"]
            ):
                mapping["debit"] = i

            # Credit column
            if "credit" not in mapping and any(
                kw in header for kw in ["credit", "deposit"]
            ):
                mapping["credit"] = i

            # Balance column
            if "balance" not in mapping and any(
                kw in header for kw in ["balance", "bal", "running"]
            ):
                mapping["balance"] = i

            # Category column
            if "category" not in mapping and "category" in header:
                mapping["category"] = i

            # Check number column
            if "check" not in mapping and "check" in header:
                mapping["check"] = i

            # Memo column (if separate from description)
            if "memo" not in mapping and "memo" in header:
                if mapping.get("description") != i:
                    mapping["memo"] = i

        # Validate minimum required columns
        if "date" not in mapping or "description" not in mapping:
            return None

        if "amount" not in mapping and (
            "debit" not in mapping or "credit" not in mapping
        ):
            return None

        return mapping

    def _parse_row(
        self,
        row: tuple[object, ...],
        mapping: dict[str, int],
        source_file: str,
        sheet_name: str,
    ) -> RawTransaction | None:
        """Parse a single row into a RawTransaction.

        Args:
            row: Excel row as tuple of values.
            mapping: Column mapping.
            source_file: Source file name.
            sheet_name: Worksheet name.

        Returns:
            RawTransaction or None.
        """
        # Get date
        date_val = self._safe_get(row, mapping.get("date"))
        if date_val is None:
            return None

        parsed_date = self._parse_date_value(date_val)
        if parsed_date is None:
            return None

        # Get description
        description = self._safe_get(row, mapping.get("description"))
        if not description:
            return None
        description = str(description).strip()

        # Get amount
        amount: Decimal | None = None
        transaction_type: TransactionType | None = None

        if "amount" in mapping:
            amount_val = self._safe_get(row, mapping["amount"])
            if amount_val is not None:
                amount = self._parse_amount_value(amount_val)
                if amount is not None:
                    transaction_type = (
                        TransactionType.CREDIT if amount >= 0 else TransactionType.DEBIT
                    )
        else:
            # Separate debit/credit columns
            debit_val = self._safe_get(row, mapping.get("debit"))
            credit_val = self._safe_get(row, mapping.get("credit"))

            debit_amt = (
                self._parse_amount_value(debit_val) if debit_val is not None else None
            )
            credit_amt = (
                self._parse_amount_value(credit_val) if credit_val is not None else None
            )

            if debit_amt is not None and debit_amt != 0:
                amount = -abs(debit_amt)
                transaction_type = TransactionType.DEBIT
            elif credit_amt is not None and credit_amt != 0:
                amount = abs(credit_amt)
                transaction_type = TransactionType.CREDIT

        if amount is None:
            return None

        # Get optional fields
        balance: Decimal | None = None
        if "balance" in mapping:
            balance_val = self._safe_get(row, mapping["balance"])
            if balance_val is not None:
                balance = self._parse_amount_value(balance_val)

        category = None
        if "category" in mapping:
            cat_val = self._safe_get(row, mapping["category"])
            if cat_val:
                category = str(cat_val).strip()

        check_number = None
        if "check" in mapping:
            check_val = self._safe_get(row, mapping["check"])
            if check_val:
                check_number = str(check_val).strip()

        memo = None
        if "memo" in mapping:
            memo_val = self._safe_get(row, mapping["memo"])
            if memo_val:
                memo = str(memo_val).strip()

        return RawTransaction(
            date=parsed_date,
            description=description,
            amount=amount,
            transaction_type=transaction_type,
            balance=balance,
            source_file=source_file,
            original_category=category,
            check_number=check_number,
            memo=memo,
            raw_data={
                "sheet_name": sheet_name,
                "row": [str(v) if v is not None else None for v in row],
                "source": "excel",
            },
        )

    def _safe_get(
        self, row: tuple[object, ...], idx: int | None
    ) -> object | None:
        """Safely get value from row.

        Args:
            row: Excel row.
            idx: Column index.

        Returns:
            Value at index or None.
        """
        if idx is None or idx < 0 or idx >= len(row):
            return None
        return row[idx]

    def _parse_date_value(self, value: object) -> date | None:
        """Parse date from Excel cell value.

        Args:
            value: Cell value (could be datetime, string, or number).

        Returns:
            Parsed date or None.
        """
        if value is None:
            return None

        # Handle datetime objects directly
        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        # Try parsing as string
        try:
            return parse_date(str(value))
        except Exception:
            return None

    def _parse_amount_value(self, value: object) -> Decimal | None:
        """Parse amount from Excel cell value.

        Args:
            value: Cell value (could be number or string).

        Returns:
            Parsed amount or None.
        """
        if value is None:
            return None

        # Handle numeric values directly
        if isinstance(value, (int, float)):
            return Decimal(str(value))

        if isinstance(value, Decimal):
            return value

        # Try parsing as string
        try:
            abs_amount, is_negative = parse_amount(str(value))
            return -abs_amount if is_negative else abs_amount
        except Exception:
            return None
